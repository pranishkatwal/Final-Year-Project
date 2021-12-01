from django.shortcuts import render
from .resources import MedicineResouce
from tablib import Dataset
from manage_medicines.models import Medicine
import openpyxl

from django.contrib.auth.decorators import login_required

@login_required
def automateFile(request):
    if request.method == 'POST':
        # print(request.FILES)
        excel_file = request.FILES["excel_file"]
        # print("############")/

        # you may put validations here to check extension or file size

        wb = openpyxl.load_workbook(excel_file)
        print(wb.sheetnames) # getting thse sheetname
        # getting a particular sheet by name out of many sheets
        worksheet = wb["Sheet1"]


        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        for row in worksheet.iter_rows():
            print("entered into for loop")
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            # print(row_data)
            excel_data.append(row_data)
        for ed in excel_data[1:]:
            temp = {}
            temp= {
                    'med_name':ed[0],
                    'price':ed[1],
                    'selling_price':ed[2],
                    'supplier':ed[3],
                    'bought_date':ed[4],
                    'expiry_date':ed[5],
                    'time_table':ed[6],
                    'quantity':ed[7] # default coz db design
                }
            Medicine.objects.create(**temp)
        return render(request,'manage_medicine.html',{'is_imported':True,'medicines':Medicine.objects.all()})


        	
        
    
    return render(request, 'automateFile.html')